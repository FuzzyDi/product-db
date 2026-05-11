"""GET/PUT /api/v1/products — каталог товаров."""
import csv
import io
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select, text, update
from sqlalchemy.ext.asyncio import AsyncSession

from product_db.db.session import get_db
from product_db.models.db import Brand, BrandAlias, Product, ProductBarcode, ProductType
from product_db.models.schemas import (
    ApiResponse,
    ProductListResponse,
    ProductResponse,
    ProductUpdateRequest,
)
from product_db.pipeline.generate import build_canonical

router = APIRouter(prefix="/products", tags=["products"])


@router.get("", response_model=ApiResponse)
async def list_products(
    status: str | None = None,
    brand_id: int | None = None,
    product_type_id: int | None = None,
    review_required: bool | None = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    q = select(Product)
    if status:
        q = q.where(Product.status == status)
    if brand_id is not None:
        q = q.where(Product.brand_id == brand_id)
    if product_type_id is not None:
        q = q.where(Product.product_type_id == product_type_id)
    if review_required is not None:
        q = q.where(Product.review_required.is_(review_required))

    total = await db.scalar(select(func.count()).select_from(q.subquery()))
    result = await db.execute(q.order_by(Product.created_at.desc()).offset(offset).limit(limit))
    items = result.scalars().all()

    # Загружаем штрихкоды для всех товаров одним запросом
    product_ids = [p.product_id for p in items]
    bc_map: dict = {}
    if product_ids:
        bc_result = await db.execute(
            select(ProductBarcode.product_id, ProductBarcode.barcode)
            .where(ProductBarcode.product_id.in_(product_ids))
        )
        for row in bc_result.all():
            bc_map.setdefault(row.product_id, []).append(row.barcode)

    responses = []
    for p in items:
        d = {k: v for k, v in p.__dict__.items() if not k.startswith('_')}
        d['barcodes'] = bc_map.get(p.product_id, [])
        responses.append(ProductResponse.model_validate(d))

    data = ProductListResponse(
        items=responses,
        total=total or 0,
        offset=offset,
        limit=limit,
    )
    return ApiResponse(data=data.model_dump())


@router.get("/search", response_model=ApiResponse)
async def search_products(
    q: str = Query(..., min_length=2),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    """Полнотекстовый + trigram поиск по name_canonical, name_normalized и штрихкоду."""
    result = await db.execute(
        text(
            """
            SELECT DISTINCT ON (p.product_id) p.*,
                   similarity(p.name_canonical, :q) AS sim
            FROM products p
            LEFT JOIN product_barcodes pb ON pb.product_id = p.product_id
            WHERE p.name_canonical % :q
               OR p.name_normalized ILIKE :like
               OR pb.barcode = :exact
               OR p.mxik_code = :exact
            ORDER BY p.product_id, sim DESC
            OFFSET :offset LIMIT :limit
            """
        ),
        {"q": q, "like": f"%{q.lower()}%", "exact": q.strip(), "offset": offset, "limit": limit},
    )
    rows = result.mappings().all()
    items = [ProductResponse.model_validate(dict(r)) for r in rows]
    return ApiResponse(data={"items": [i.model_dump() for i in items], "count": len(items)})


@router.get("/export/xlsx")
async def export_xlsx(
    status: str = Query("certified"),
    db: AsyncSession = Depends(get_db),
):
    """Выгрузка товаров в XLSX."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    q = select(Product)
    if status:
        q = q.where(Product.status == status)
    q = q.order_by(Product.created_at.asc())

    result = await db.execute(q)
    products = result.scalars().all()

    product_ids = [p.product_id for p in products]
    bc_map: dict = {}
    if product_ids:
        bc_result = await db.execute(
            select(ProductBarcode.product_id, ProductBarcode.barcode)
            .where(ProductBarcode.product_id.in_(product_ids))
        )
        for row in bc_result.all():
            bc_map.setdefault(row.product_id, []).append(row.barcode)

    COLUMNS = [
        ("product_id",        "ID"),
        ("status",            "Статус"),
        ("name_raw",          "Сырое название"),
        ("name_canonical",    "Канонич. название"),
        ("name_pos",          "POS (≤20)"),
        ("name_receipt",      "Чек (≤40)"),
        ("brand_name",        "Бренд"),
        ("variant",           "Вариант"),
        ("quantity_value",    "Кол-во"),
        ("quantity_unit",     "Единица"),
        ("package_code",      "Упаковка"),
        ("mxik_code",         "ИКПУ"),
        ("mxik_package_code", "Пакейдж код"),
        ("label_required",    "Маркируемый"),
        ("label_for_check",   "Марка на чек"),
        ("cash_sale",         "Наличные"),
        ("confidence_score",  "Confidence"),
        ("completeness_score","Completeness"),
        ("barcodes",          "Штрихкоды"),
        ("created_at",        "Создан"),
        ("updated_at",        "Обновлён"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = status

    # Заголовок
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    for col_idx, (_, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    # Данные
    for row_idx, p in enumerate(products, 2):
        for col_idx, (field, _) in enumerate(COLUMNS, 1):
            if field == "product_id":
                val = str(p.product_id)
            elif field == "barcodes":
                val = ";".join(bc_map.get(p.product_id, []))
            elif field in ("created_at", "updated_at"):
                dt = getattr(p, field, None)
                val = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
            else:
                raw = getattr(p, field, None)
                val = float(raw) if raw is not None and field in (
                    "quantity_value", "confidence_score", "completeness_score"
                ) else raw
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Ширина колонок
    col_widths = [36, 12, 40, 40, 22, 42, 20, 20, 8, 8, 12, 18, 14, 12, 12, 10, 12, 12, 30, 18, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"products_{status}.xlsx"
    return StreamingResponse(
        iter([output.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{product_id}", response_model=ApiResponse)
async def get_product(product_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Product).where(Product.product_id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Not found")
    bc_result = await db.execute(
        select(ProductBarcode.barcode).where(ProductBarcode.product_id == product_id)
    )
    barcodes = [row.barcode for row in bc_result.all()]
    d = {k: v for k, v in product.__dict__.items() if not k.startswith('_')}
    d['barcodes'] = barcodes
    return ApiResponse(data=ProductResponse.model_validate(d).model_dump())


@router.put("/{product_id}", response_model=ApiResponse)
async def update_product(
    product_id: uuid.UUID,
    req: ProductUpdateRequest,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Product).where(Product.product_id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Not found")

    values = req.model_dump(exclude_none=True)

    # Обрабатываем brand_name отдельно — ищем или создаём бренд
    brand_name = values.pop("brand_name", None)
    if brand_name:
        canonical = brand_name.strip().upper()
        brand = await db.scalar(select(Brand).where(Brand.name_canonical == canonical))
        if not brand:
            brand = Brand(name_canonical=canonical)
            db.add(brand)
            await db.flush()
            db.add(BrandAlias(brand_id=brand.id, alias=canonical, source="operator"))
        values["brand_id"] = brand.id
        values["brand_name"] = canonical
        values["issues"] = func.array_remove(Product.issues, "MISSING_BRAND")

        # Перестраиваем canonical если бренда раньше не было
        if not product.brand_id:
            product_type_name = None
            if product.product_type_id:
                product_type_name = await db.scalar(
                    select(ProductType.name_ru).where(ProductType.id == product.product_type_id)
                )
            new_canonical = build_canonical(
                product_type=product_type_name,
                brand=canonical,
                subbrand=None,
                variant=product.variant if hasattr(product, 'variant') else None,
                quantity_value=product.quantity_value,
                quantity_unit=product.quantity_unit,
                package_code=product.package_code,
                name_raw=product.name_raw,
            )
            # Не перезаписываем если оператор сам задал название
            if "name_canonical" not in values:
                values["name_canonical"] = new_canonical
                values["name_pos"] = new_canonical[:20]
                values["name_receipt"] = new_canonical[:40]

    # При смене типа товара — убираем MISSING_PRODUCT_TYPE и пересчитываем canonical
    new_product_type_id = values.get("product_type_id")
    if new_product_type_id and not product.product_type_id:
        values["issues"] = func.array_remove(
            values.get("issues", Product.issues), "MISSING_PRODUCT_TYPE"
        )
        if "name_canonical" not in values:
            pt_name = await db.scalar(
                select(ProductType.name_ru).where(ProductType.id == new_product_type_id)
            )
            brand = values.get("brand_name") or product.brand_name
            new_canonical = build_canonical(
                product_type=pt_name,
                brand=brand,
                subbrand=None,
                variant=None,
                quantity_value=product.quantity_value,
                quantity_unit=product.quantity_unit,
                package_code=product.package_code,
                name_raw=product.name_raw,
            )
            values["name_canonical"] = new_canonical
            values["name_pos"] = new_canonical[:20]
            values["name_receipt"] = new_canonical[:40]

    if values:
        await db.execute(update(Product).where(Product.product_id == product_id).values(**values))
        await db.commit()
        await db.refresh(product)

    d = {k: v for k, v in product.__dict__.items() if not k.startswith('_')}
    bc_result = await db.execute(
        select(ProductBarcode.barcode).where(ProductBarcode.product_id == product_id)
    )
    d['barcodes'] = [row.barcode for row in bc_result.all()]
    return ApiResponse(data=ProductResponse.model_validate(d).model_dump())
